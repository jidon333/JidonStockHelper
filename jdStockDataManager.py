

import FinanceDataReader as fdr

import yahooquery as yq 
from yahooquery import Ticker

import pickle
import math

import numpy as np
import pandas as pd
import pandas_market_calendars as mcal
from pandas import Timestamp, DatetimeIndex


import os
import json
import datetime as dt
import time
import random
import concurrent.futures
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import PatternFill, Font, Color

import logging

from jdGlobal import get_yes_no_input
from jdGlobal import (
    DATA_FOLDER,
    METADATA_FOLDER,
    FILTERED_STOCKS_FOLDER,
    SCREENSHOT_FOLDER,
    PROFILES_FOLDER,
    sync_fail_ticker_list,
    exception_ticker_list
)

from jd_io_utils import (
    ensure_directories_exist,
    load_csv_with_date_index,
    save_df_to_csv,
    save_to_json,
    load_from_json,
    save_pickle,
    load_pickle
)

from jdDataGetter import JdDataGetter

import jdIndicator as jdi


# ──────────────────────────────────────────────────────────────────────────────
# 📌 QUICK REFERENCE — Date-Time & Time-Zone Helpers
#
# 1️⃣ 객체 종류
# ────────────────────────────────────────────────────────────────────────────
# • pd.Timestamp
#     단일 시점(“time point”)
#       예) pd.Timestamp("2025-05-06 09:30")
#
# • pd.DatetimeIndex
#     날짜·시간 값이 “여러 개” 모여 있는 인덱스(일종의 리스트)
#       예) pd.date_range("2025-05-01", periods=3)  # 5/1, 5/2, 5/3 세 날짜가 들어감
#
# 2️⃣ tz-naive  vs  tz-aware
# ────────────────────────────────────────────────────────────────────────────
# • tz-naive  : 타임존 정보(tzinfo)가 없음
#       예) Timestamp('2025-05-06 09:30')
#
# • tz-aware  : 타임존 정보가 붙어 있음 → ‘어느 지역 시간’인지 명확
#       예) Timestamp('2025-05-06 09:30:00+00:00', tz='UTC')
#       예) Timestamp('2025-05-06 09:30:00-04:00', tz='America/New_York')
#
#   ⚠️ pandas는 tz가 다른 두 객체를 직접 비교·연산하면
#      TypeError(“Cannot compare tz-naive and tz-aware …”) 를 던진다.
#
# 3️⃣ 타임존 처리 메서드
# ────────────────────────────────────────────────────────────────────────────
# • tz_localize("UTC")
#       tz-naive  ➜  tz-aware(UTC)  │  *라벨만 부여* (값은 그대로)
#       예) "2025-05-06 09:30"  ➜  "2025-05-06 09:30+00:00"
#
# • tz_convert("UTC")
#       tz-aware ➜  다른 타임존(UTC) │  *값을 변환*
#       예) "2025-05-06 09:30-04:00(NY)" ➜ "2025-05-06 13:30+00:00"
#
# 4️⃣ 헬퍼 함수 (아래 정의)
# ────────────────────────────────────────────────────────────────────────────
# • to_utc_ts(ts: Timestamp)        →  항상 tz-aware(UTC) Timestamp 반환
# • to_utc_idx(idx: DatetimeIndex)  →  항상 tz-aware(UTC) DatetimeIndex 반환
#
#   사용 예)
#       naive_ts  = pd.Timestamp("2025-05-06 09:30")                   # naive
#       aware_ts  = pd.Timestamp("2025-05-06 09:30", tz="US/Eastern")  # aware
#
#       utc_ts1 = to_utc_ts(naive_ts)   # 2025-05-06 09:30+00:00
#       utc_ts2 = to_utc_ts(aware_ts)   # 2025-05-06 13:30+00:00
#
#       idx     = pd.date_range("2025-05-01", periods=3)              # naive idx
#       utc_idx = to_utc_idx(idx)   # DatetimeIndex(['2025-05-01 00:00+00:00', …])
#
#   이렇게 변환해 두면 서로 간 비교·필터링·머지 시 안전하다.
# ──────────────────────────────────────────────────────────────────────────────

def to_utc_ts(ts: Timestamp) -> Timestamp:
    """Timestamp → tz-aware(UTC) 로 강제 변환."""
    return ts.tz_localize("UTC") if ts.tz is None else ts.tz_convert("UTC")

def to_utc_idx(idx: DatetimeIndex) -> DatetimeIndex:
    """DatetimeIndex → tz-aware(UTC) 로 강제 변환."""
    return idx.tz_localize("UTC") if idx.tz is None else idx.tz_convert("UTC")



# ----------------------------
# 미국 시장 달력
# ----------------------------
nyse = mcal.get_calendar('NYSE')




# for test
stockIterateLimit = 99999






class JdStockDataManager:
    def __init__(self):
        """
        생성자에서 ensure_directories_exist()를 호출해
        모든 폴더(DATA_FOLDER, METADATA_FOLDER 등)가 존재하는지 확인하고,
        없으면 생성.
        """

        ensure_directories_exist()

        # S&P500 지수 데이터(기본값)
        self.us500_data = fdr.DataReader('US500')
        self.stock_GICS_df = pd.DataFrame()

        self.long_term_industry_rank_df = pd.DataFrame()
        self.short_term_industry_rank_df = pd.DataFrame()
        self.atrs_ranking_df = pd.DataFrame()

        # Cache
        self.reset_caches()


        # DataGetter 인스턴스를 미리 생성하여 멤버 변수로 저장
        self.data_getter = JdDataGetter()

    def reset_caches(self):
        self.cache_StockListFromLocalCsv = pd.DataFrame()
        self.cache_getStockDatasFromCsv_out_tickers = None
        self.cache_getStockDatasFromCsv_out_stock_datas_dic = None



    # ------------------------------------------------
    # [ Private ]: 내부 함수들
    # ------------------------------------------------

    def _get_csv_names(self):
        """
        data 폴더 내 모든 .csv 파일명을 ticker 리스트로 반환
        예: SPY.csv -> 'SPY'
        """
        csv_names =[os.path.splitext(f)[0] for f in os.listdir(DATA_FOLDER) if f.endswith('.csv')]
        return csv_names



    def _CookIndexData(self, index_data, n = 14):
        """
        S&P500 (index_data)에 ATR, TC, ATC 컬럼을 추가
        """
        index_new_data = index_data

        # TR(True Range) 계산
        tr = jdi.true_range(index_new_data)

        # ATR(Average True Range)
        atr = jdi.atr(index_new_data)
        index_new_data['ATR'] = atr

        # TC(True Change) 계산
        tc = (index_new_data['Close'] - index_new_data['Close'].shift(1)) / atr
        index_new_data['TC'] = tc

        # ATC(Average True Change) 계산
        atc = tc.rolling(n).mean()
        index_new_data['ATC'] = atc

        return index_new_data

    def _CookStockData(self, stock_data : pd.DataFrame):
        """
        종목별 주가 데이터를 받아서
        RS, MA, ATR, TRS, ATRS 등 각종 지표를 추가
        """
        new_data = stock_data

        try:
            # MRS 계산
            n = 20
            rs = (stock_data['Close'] / self.us500_data['Close']) * 100
            rs_ma = rs.rolling(n).mean()
            mrs = ((rs / rs_ma) - 1) * 100

            # MRS를 주식 데이터에 추가
            new_data['RS'] = mrs

            # 50MA
            new_data['50MA'] = jdi.sma(stock_data['Close'], 50)

            # 150MA
            new_data['150MA'] = jdi.sma(stock_data['Close'], 150)

            # 200MA
            new_data['200MA'] = jdi.sma(stock_data['Close'], 200)

            # 150MA Slope
            ma_diff = stock_data['150MA'].diff()
            new_data['MA150_Slope'] = ma_diff / 2

            # 200MA Slope
            ma_diff = stock_data['200MA'].diff()
            new_data['MA200_Slope'] = ma_diff / 2


            # TR 계산
            new_data['TR'] = jdi.true_range(stock_data)

            # ADR(%)
            new_data['ADR'] = jdi.adr(stock_data, 20)

            # ATR 계산
            atr = jdi.atr(stock_data, 14)
            new_data['ATR'] = atr

            # TC(True Change) 계산
            tc = (stock_data['Close'] - stock_data['Close'].shift(1)) / atr
            new_data['TC'] = tc

            # ATC(Average True Change) 계산
            atc = tc.rolling(n).mean()
            new_data['ATC'] = atc

            new_index_data = self._CookIndexData(self.us500_data, 14)
        
            # TRS(True Relative Strength)
            sp500_tc = new_index_data['TC']
            stock_tc = tc
            trs = stock_tc - sp500_tc
            new_data['TRS'] = trs

            # ATRS (14 days Average True Relative Strength)
            atrs = trs.rolling(n).mean()
            new_data['ATRS'] = atrs

            atrs_exp = trs.ewm(span=14, adjust=False).mean()
            new_data['ATRS_Exp'] = atrs_exp

            n = 150  # 이동평균 윈도우 크기
            # ATRS150 (150 days Average True Relative Strength)

            if len(new_data) < n:
                atrs150 = trs.rolling(len(new_data)).mean()
                new_data['ATRS150'] = atrs150
            else:
                atrs150 = trs.rolling(n).mean()
                new_data['ATRS150'] = atrs150

            # ATRS150_Exp 
            span = len(new_data) if len(new_data) < n else n
            new_data["ATRS150_Exp"] = jdi.ema(trs, span)


            new_data = new_data.reindex(columns=['Symbol', 'Name', 'Industry',
                                                 'Open', 'High', 'Low', 'Close', 'Adj Close',
                                                 'Volume', 'RS','50MA', '150MA', '200MA',
                                                 'MA150_Slope', 'MA200_Slope', 
                                                 'ADR', 'TR', 'ATR', 'TC', 'ATC', 'TRS', 'ATRS', 'ATRS_Exp', 'ATRS150', 'ATRS150_Exp',
                                                 'IsOriginData_NaN'])
            

            new_data = new_data.round(5)


        except Exception as e:
            print(e)
            raise

        return new_data
   
    def _export_datas_to_csv(self, data_dic):
        """
        딕셔너리 {ticker: DataFrame}을 각각 CSV로 저장.
        """
        i = 0
        total_len = len(data_dic.items())
        for ticker, df in data_dic.items():
            save_df_to_csv(df, ticker)
            i = i+1
            print(f"{ticker}.csv is saved! {i}/{total_len}")



    
    def _getATRCondition_df(self, stock_list, ticker):
        """
        ATR 기반으로 condition A/B를 계산.
        """
        try:
            save_path = os.path.join(DATA_FOLDER, f"{ticker}.csv")
            data = pd.read_csv(save_path)
            data.set_index('Date', inplace=True)

            ATR = data['ATR']
            volume_ma50 = data['Volume'].rolling(window=50).mean()
            open = data['Open']
            close = data['Close']
            diff = close - open

            bIsVolumeEnough = (volume_ma50 >= 2000000) & (close >= 10)
            conditionA = bIsVolumeEnough & (diff > (1.0 * ATR))
            conditionB = bIsVolumeEnough & (diff < (-1.5 * ATR))
     
        except Exception as e:
                print(f"[_getATRCondition_df] An error occurred: {e} for ticker {ticker}")
                name = stock_list.loc[stock_list['Symbol'] == ticker, 'Name'].values[0]
                exception_ticker_list[ticker] = name
                conditionA = pd.Series()
                conditionB = pd.Series()
        return conditionA, conditionB
    

    def _getUpDownConditions_df(self, stock_list):
        l_A = []
        l_B = []
        for ticker in stock_list['Symbol']:
            conditionA, conditionB = self._getATRCondition_df(stock_list, ticker)
            l_A.append(conditionA)
            l_B.append(conditionB)
        
        all_conditions_A = pd.concat(l_A, axis=1, sort=True)
        all_conditions_B = pd.concat(l_B, axis=1, sort=True)
        all_conditions_A.columns = all_conditions_A.columns.to_list()
        all_conditions_B.columns = all_conditions_B.columns.to_list()

        daily_changes = pd.DataFrame(index=all_conditions_A.index, columns=['conditionA', 'conditionB'])
        daily_changes['conditionA'] = all_conditions_A.sum(axis=1)  # 조건 A를 만족하는 종목 수
        daily_changes['conditionB'] = all_conditions_B.sum(axis=1)  # 조건 B를 만족하는 종목 수
        daily_changes['sum'] = daily_changes['conditionA'] - daily_changes['conditionB']  # 조건 A와 B의 차이
        daily_changes['ma200_changes'] = daily_changes['sum'].rolling(200).mean()  # 150일 이동 평균
        daily_changes['ma50_changes'] = daily_changes['sum'].rolling(50).mean()  # 150일 이동 평균
        daily_changes['ma20_changes'] = daily_changes['sum'].rolling(20).mean()  # 150일 이동 평균

        daily_changes['conditionA'] = pd.to_numeric(daily_changes['conditionA'], errors='coerce').astype('Int64')
        daily_changes['conditionB'] = pd.to_numeric(daily_changes['conditionB'], errors='coerce').astype('Int64')
        daily_changes['sum']        = pd.to_numeric(daily_changes['sum'],        errors='coerce').astype('Int64')
        daily_changes['ma200_changes'] = pd.to_numeric(daily_changes['ma200_changes'], errors='coerce').astype('float64')
        daily_changes['ma50_changes']  = pd.to_numeric(daily_changes['ma50_changes'],  errors='coerce').astype('float64')
        daily_changes['ma20_changes']  = pd.to_numeric(daily_changes['ma20_changes'],  errors='coerce').astype('float64')


        # ✅ 인덱스를 날짜로 확정 + 이름을 'Date'로 지정 → CSV 첫 열 헤더가 항상 'Date'
        daily_changes.index = pd.to_datetime(daily_changes.index)
        daily_changes.index.name = 'Date'

        return daily_changes



    def _getCloseChanges_df(self, stock_list, ticker):
        try:
            save_path = os.path.join(DATA_FOLDER, f"{ticker}.csv")
            data = pd.read_csv(save_path)
            data.set_index('Date', inplace=True)
            returns = (data['Close'] - data['Close'].shift(1)) / data['Close'].shift(1)
        except Exception as e:
            print(f"[_getCloseChanges_df] An error occurred: {e} for ticker {ticker}")
            name = stock_list.loc[stock_list['Symbol'] == ticker, 'Name'].values[0]
            exception_ticker_list[ticker] = name
            returns = pd.Series()


        return returns

    def _getUpDownChanges_df(self, stock_list):
        # 모든 종목에 대한 전일 대비 수익률 계산
        l = [self._getCloseChanges_df(stock_list, ticker) for ticker in stock_list['Symbol']]
        all_returns = pd.concat(l, axis=1, sort=True)
        all_returns.columns = all_returns.columns.to_list()

        # all_returns의 각 행은 날짜, 각 열은 종목을 의미.
        # sum의 axis = 0은 모든 행을 더하고, axis = 1은 모든 열을 더한다.
        # (all_returns > 0)으로 all_returns의 모든 값을 True or False로 변경하고
        # sum 함수를 이용해 모든 열을 더해 상승 종목 수와 하락 종목 수를 구한다.
        daily_changes = pd.DataFrame(index=all_returns.index, columns=['up', 'down'])
        daily_changes['up'] = (all_returns > 0).sum(axis=1)
        daily_changes['down'] = (all_returns < 0).sum(axis=1)
        daily_changes['sum'] = daily_changes['up'] - daily_changes['down']
        daily_changes['ma150_changes'] = daily_changes['sum'].rolling(150).mean()

        # ✅ 숫자 dtype 고정 (엑셀에서 텍스트 오인 방지)
        daily_changes['up']   = pd.to_numeric(daily_changes['up'],   errors='coerce').astype('Int64')
        daily_changes['down'] = pd.to_numeric(daily_changes['down'], errors='coerce').astype('Int64')
        daily_changes['sum']  = pd.to_numeric(daily_changes['sum'],  errors='coerce').astype('Int64')
        daily_changes['ma150_changes'] = pd.to_numeric(daily_changes['ma150_changes'], errors='coerce').astype('float64')

        # ✅ 인덱스를 날짜형으로 확정 + 이름을 'Date'로 지정
        daily_changes.index = pd.to_datetime(daily_changes.index)
        daily_changes.index.name = 'Date'   # ← 이 한 줄로 CSV 첫 열 헤더가 항상 'Date'


        return daily_changes

    # ------------------------------------------------
    # [ Public Methods ]
    # ------------------------------------------------


    def get_local_stock_list(self):
        return self.data_getter.get_local_stock_list()
    

    def cook_ATR_Expansion_Counts(self, daysNum = 365*5):
        stock_list = self.get_local_stock_list()
        up_down_condition_df = self._getUpDownConditions_df(stock_list)
        up_down_condition_df.to_csv(os.path.join(METADATA_FOLDER, 'ATR_Expansion_Counts.csv'))
    

    def cookUpDownDatas(self, daysNum = 365*5):

        # S&P 500 지수의 모든 종목에 대해 매일 상승/하락한 종목 수 계산
        nyse_list = self.data_getter.get_fdr_stock_list('NYSE', daysNum)
        nyse_list = nyse_list[nyse_list['Symbol'].isin(self._get_csv_names())]

        nasdaq_list = self.data_getter.get_fdr_stock_list('NASDAQ', daysNum)
        nasdaq_list = nasdaq_list[nasdaq_list['Symbol'].isin(self._get_csv_names())]

        sp500_list = self.data_getter.get_fdr_stock_list('S&P500', daysNum)
        sp500_list = sp500_list[sp500_list['Symbol'].isin(self._get_csv_names())]


        # 미국 주식시장의 거래일 가져오기
        nyse = mcal.get_calendar('NYSE')
        trading_days = nyse.schedule(start_date=dt.date.today() - dt.timedelta(days=daysNum), end_date=dt.date.today()).index
        valid_start_date = trading_days[0]
        valid_end_date = trading_days[-1]

        daily_changes_nyse_df = self._getUpDownChanges_df(nyse_list)
        daily_changes_nyse_df.to_csv(os.path.join(METADATA_FOLDER, 'up_down_nyse.csv'))

        daily_changes_nasdaq_df = self._getUpDownChanges_df(nasdaq_list)
        daily_changes_nasdaq_df.to_csv(os.path.join(METADATA_FOLDER, 'up_down_nasdaq.csv'))

        daily_changes_sp500_df = self._getUpDownChanges_df(sp500_list)
        daily_changes_sp500_df.to_csv(os.path.join(METADATA_FOLDER, 'up_down_sp500.csv'))

        with open("up_down_exception.json", "w") as outfile:
                json.dump(exception_ticker_list, outfile, indent = 4)

        return daily_changes_nyse_df, daily_changes_nasdaq_df, daily_changes_sp500_df
    
    
    def _get_last_completed_trading_day(self, cal: mcal.MarketCalendar) -> pd.Timestamp:
        """
        NYSE 기준 ‘이미 마감된’ 가장 최근 거래일을 tz-aware(UTC) Timestamp 로 반환.
            · 오늘이 휴장       → 직전 거래일
            · 오늘이 장중       → 직전 거래일
            · 오늘이 장마감 이후 → 오늘
        """
        now_et = dt.datetime.now(ZoneInfo("America/New_York"))
        today_sched = cal.schedule(now_et.date(), now_et.date())

        # 오늘이 거래일이고 장이 완전히 끝난 경우 → 오늘
        if (not today_sched.empty) and now_et >= today_sched.iloc[0]["market_close"]:
            return to_utc_ts(today_sched.index[0])

        # 그 외(휴일·장중) → 직전 거래일
        prev_sched = cal.schedule(
            start_date=now_et.date() - dt.timedelta(days=10),
            end_date=now_et.date() - dt.timedelta(days=1)
        )
        return to_utc_ts(prev_sched.index[-1])

    def cook_filter_count_data(
        self,
        filter_func,
        fileName: str,
        daysNum: int = 365,
        bAccumulateToExistingData: bool = True,
    ):
        """
        · filter_func(stock_dic: dict[str, DataFrame], offset: int) -> list[str]
            offset = -1 → 어제, -2 → 그저께 …
        · fileName : '{fileName}.csv' 로 저장
        · daysNum  : 마지막 완료 세션으로부터 N 일 전까지 계산
        """
        # 1) 데이터 준비 --------------------------------------------------------
        stock_list        = self.get_local_stock_list()
        stock_data_len    = 365 * 5
        stock_dic         = self.get_stock_datas_from_csv(stock_list, stock_data_len, bAccumulateToExistingData)

        # 2) 거래일 & 마지막 완료 세션 -----------------------------------------
        nyse          = mcal.get_calendar("NYSE")
        trading_days  = nyse.valid_days(
        start_date=dt.date.today() - dt.timedelta(days=daysNum),
        end_date  =dt.date.today(),
        )

        trading_days  = to_utc_idx(trading_days) # tz-aware(UTC)

        last_completed = self._get_last_completed_trading_day(nyse)
        completed_days = trading_days[trading_days <= last_completed]

        # 3) 필터 루프 ----------------------------------------------------------
        days, cnts = [], []
        for off, day in enumerate(completed_days[::-1], start=1):
            tickers = filter_func(stock_dic, -off)
            days.append(day)
            cnts.append(len(tickers))
            print(f"{fileName}: {day.date()} → {len(tickers)}")

        result_df = (pd.DataFrame({"Date": days[::-1], "Count": cnts[::-1]})
                    .set_index("Date"))                       # tz-aware(UTC)

        # 4) 파일 병합 ----------------------------------------------------------
        save_path = os.path.join(METADATA_FOLDER, f"{fileName}.csv")

        if bAccumulateToExistingData and os.path.exists(save_path):
            local = pd.read_csv(save_path, index_col="Date", parse_dates=["Date"])
            # 과거 CSV는 tz-naive → UTC 로 통일
            local.index = to_utc_idx(local.index)
            result_df   = pd.concat([local[~local.index.isin(result_df.index)],
                                    result_df])

        # 5) 저장: 사람이 보기 좋도록 tz 정보 제거 -----------------------------
        export_df = result_df.copy()
        if export_df.index.tz is not None:           # tz-aware → tz-naive
            export_df.index = export_df.index.tz_localize(None)

        export_df.to_csv(save_path, encoding="utf-8-sig")


    def get_count_data_from_csv(self, fileName : str, daysNum = 365*2):
            """ 
            fileName: {fileName}_Counts.csv 
            """
            # ------------ nyse -------------------
            data_path = os.path.join(METADATA_FOLDER, f"{fileName}_Counts.csv")
            if not os.path.exists(data_path):
                print(f"No file: {data_path}")
                return pd.DataFrame()
            
            data = pd.read_csv(data_path)

            # 문자열을 datetime 객체로 변경
            data['Date'] = pd.to_datetime(data['Date'])

            # Date 행을 인덱스로 설정
            data.set_index('Date', inplace=True)

            # 미국 주식시장의 거래일 가져오기
            trading_days = nyse.schedule(start_date=dt.date.today() - dt.timedelta(days=daysNum), end_date=dt.date.today()).index
            startDay = trading_days[0].date()
            endDay = min(trading_days[-1], data.index[-1]).date()

            # 시작일부터 종료일까지 가져오기
            cnt_data = data[startDay:endDay]
            return cnt_data

    def download_stock_datas_from_web(self, days_num=365*5, exclude_not_in_local_csv=True):
        """
        DataGetter에서 웹 데이터만 가져와 병합 결과 반환 -> 여기서 Cook + CSV 저장
        """

        # override 경고 및 confirm
        yes_no = get_yes_no_input("It will override all your local .csv files. Continue? (y/n)")
        if not yes_no:
            return

        merged_data_dic = self.data_getter.download_stock_datas_from_web(days_num, exclude_not_in_local_csv)
        cooked_data_dic = {}
        i = 0
        total = len(merged_data_dic)


        for ticker, df in merged_data_dic.items():
            if df.empty:
                sync_fail_ticker_list.append(ticker)
                continue

            try:
                cooked_df = self._CookStockData(df)
                cooked_data_dic[ticker] = cooked_df
                i += 1
                print(f"[download_stock_datas_from_web] {ticker} cooked. {i}/{total}")
            except Exception as e:
                print(f"Error cooking {ticker}: {e}")
                exception_ticker_list[ticker] = str(e)

        # 최종 CSV로 저장
        for t, cdf in cooked_data_dic.items():
            save_df_to_csv(cdf, t, data_dir=DATA_FOLDER)

        # 실패 목록
        with open('download_fail_list.txt', 'w') as f:
            for tk in sync_fail_ticker_list:
                f.write(tk + '\n')

        print("[download_stock_datas_from_web] Done override.")

    # cooking 공식이 변하는 경우 로컬 데이터를 업데이트하기 위해 호출
    def cookLocalStockData(self, bUseLocalCache = False):
        print("-------------------cookLocalStockData-----------------\n ") 

        all_list = self.get_local_stock_list()
        stock_datas_from_csv = self.get_stock_datas_from_csv(all_list, 365*6, bUseLocalCache)
        tickers = list(stock_datas_from_csv.keys())


        cooked_data_dic = {}

        for ticker in tickers:
            csvData = stock_datas_from_csv.get(ticker, pd.DataFrame())

            if csvData.empty:
                sync_fail_ticker_list.append(ticker)
                continue

            cookedData = self._CookStockData(csvData)
            cooked_data_dic[ticker] = cookedData

            print(ticker, ' cooked!')

        self._export_datas_to_csv(cooked_data_dic)

    def sync_csv_from_web(self, daysNum = 14):
        """
        - DataGetter.sync_csv_from_web으로부터 병합된 “원본+웹” merged_data_dic을 받음
        - 여기서 Cook 후 CSV 저장
        """
        print("[JdStockDataManager] sync_csv_from_web start...")
        merged_data_dic = self.data_getter.sync_csv_from_web(daysNum)

        cooked_data_dic = {}
        i = 0
        total = len(merged_data_dic.keys())

        for ticker, merged_df in merged_data_dic.items():
            if merged_df.empty:
                sync_fail_ticker_list.append(ticker)
                continue

        
            # [Cook]
            try:
                cooked_df = self._CookStockData(merged_df)
                cooked_data_dic[ticker] = cooked_df
                i += 1
                logging.info(f"[sync_csv_from_web] {ticker} cooked. {i}/{total}")
        
            except Exception as e:
                logging.info(f"Error cooking {ticker}: {e}")
                exception_ticker_list[ticker] = str(e)

        # 이제 최종 CSV 저장
        for ticker, cooked_df in cooked_data_dic.items():
            save_df_to_csv(cooked_df, ticker, data_dir=DATA_FOLDER)

        # 실패 목록 기록
        with open('sync_fail_list.txt', 'w') as f:
            for tk in sync_fail_ticker_list:
                f.write(tk + '\n')

        print("[sync_csv_from_web] Done.")

    def getUpDownDataFromCsv(self, daysNum = 365*2):
        updown_nyse = pd.DataFrame()
        updown_nasdaq = pd.DataFrame()
        updown_sp500 = pd.DataFrame()

        # ------------ nyse -------------------
        nyse_file_path = os.path.join(METADATA_FOLDER, "up_down_nyse.csv")
        data = pd.read_csv(nyse_file_path)

        # 문자열을 datetime 객체로 변경
        data['Date'] = pd.to_datetime(data['Date'])

        # Date 행을 인덱스로 설정
        data.set_index('Date', inplace=True)


        # 미국 주식시장의 거래일 가져오기
        trading_days = nyse.schedule(start_date=dt.date.today() - dt.timedelta(days=daysNum), end_date=dt.date.today()).index
        startDay = trading_days[0]
        endDay = min(trading_days[-1], data.index[-1])

        # 시작일부터 종료일까지 가져오기
        data = data[startDay:endDay]
        updown_nyse = data


        # ------------ nasdaq -------------------
        nasdaq_file_path = os.path.join(METADATA_FOLDER, "up_down_nasdaq.csv")
        data = pd.read_csv(nasdaq_file_path)

        # 문자열을 datetime 객체로 변경
        data['Date'] = pd.to_datetime(data['Date'])

        # Date 행을 인덱스로 설정
        data.set_index('Date', inplace=True)
        # 시작일부터 종료일까지 가져오기
        data = data[startDay:endDay]
        updown_nasdaq = data

        # ------------ sp500 -------------------
        sp500_file_path = os.path.join(METADATA_FOLDER, "up_down_sp500.csv")
        data = pd.read_csv(sp500_file_path)

        # 문자열을 datetime 객체로 변경
        data['Date'] = pd.to_datetime(data['Date'])

        # Date 행을 인덱스로 설정
        data.set_index('Date', inplace=True)
        # 시작일부터 종료일까지 가져오기
        data = data[startDay:endDay]
        updown_sp500 = data

        return updown_nyse, updown_nasdaq, updown_sp500

    def get_stock_datas_from_csv(self, stock_list : pd.DataFrame, daysNum = 365*5, bUseCacheData = False):
        return self.data_getter.get_stock_datas_from_csv(stock_list, daysNum, bUseCacheData)


    def remove_acquisition_tickers(self):
        all_list = self.get_local_stock_list()

        stock_datas_from_csv = self.get_stock_datas_from_csv(all_list)
        tickers = list(stock_datas_from_csv.keys())


        removeTargetTickers = []

        for ticker in tickers:
            data = stock_datas_from_csv[ticker]
            name = data['Name'].iloc[-1].lower()
            try:
                industry = data['Industry'].iloc[-1].lower()
            except Exception as e:
                removeTargetTickers.append(ticker)
                print(e)
                continue
                  
            if pd.isna(name) or pd.isna(industry):
                removeTargetTickers.append(ticker)
                continue
            if 'acquisition' in name or '기타 금융업' in industry:
                removeTargetTickers.append(ticker)
            if 'acquisition' in name or '투자 지주 회사' in industry:
                removeTargetTickers.append(ticker)


        for ticker in removeTargetTickers:
            file_path = os.path.join(DATA_FOLDER, ticker + '.csv')
            if os.path.exists(file_path):
                os.remove(file_path)
                print(file_path, 'is removed from local directory!')

    def cook_Nday_ATRS150_exp(self, N=150):
        all_list = self.get_local_stock_list()

        propertyName = 'ATRS150_Exp'

        stock_datas_from_csv = self.get_stock_datas_from_csv(all_list)
        tickers = list(stock_datas_from_csv.keys())

        date_list = None  # 변수 초기화

        atrs_dict = {}
        for ticker in tickers:
            data = stock_datas_from_csv[ticker]
            atrs_list = data[propertyName].iloc[-N:].tolist() # 최근 N일 동안의 ATRS150 값만 가져오기=
            atrs_list = [x if not math.isnan(x) else -1 for x in atrs_list] # NaN의 경우 -1로 대체
            while len(atrs_list) < N:
                atrs_list.insert(0, -1) # 리스트앞에 -1을 추가하여 과거 NaN 데이터를 -1로 치환
            if pd.notna(atrs_list).all() and len(atrs_list) == N: # ATRS150 값이 모두 유효한 경우에만 추가
                atrs_dict[ticker] = atrs_list
                if date_list is None:  # 처음으로 유효한 atrs_list를 발견하면 날짜 정보를 가져옴
                    date_list = data.index[-N:].strftime('%Y-%m-%d').tolist()

        
        atrs_df = pd.DataFrame.from_dict(atrs_dict)
        atrs_df['Date'] = date_list
        atrs_df = atrs_df.set_index('Date')
        atrs_df = atrs_df.T # transpose

        save_path = os.path.join(METADATA_FOLDER, f'{N}day_{propertyName}.csv')
        atrs_df.to_csv(save_path, encoding='utf-8-sig', index_label='Symbol')

        return atrs_df

    def cook_ATRS150_exp_Ranks(self, N = 150):

        propertyName = 'ATRS150_Exp'

        csv_path = os.path.join(METADATA_FOLDER, f'{N}day_{propertyName}.csv')
        data = pd.read_csv(csv_path)
        data = data.set_index('Symbol')

        rank_df = data.rank(axis=0, ascending=False, method='dense')

        rank_df = rank_df

        save_path = os.path.join(METADATA_FOLDER, f'{propertyName}_Ranking.csv')
        rank_df.to_csv(save_path, encoding='utf-8-sig', index_label='Symbol')

    def get_ATRS150_exp_Ranks_Normalized(self, Symbol):

        propertyName = 'ATRS150_Exp'

        try:
            rank_df = self.get_ATRS_Ranking_df()
            serise_rankChanges = rank_df.loc[Symbol]

            max_value = len(rank_df)

            # normalize ranking [0, 1] so that value '1' is most strong ranking.
            serise_rankChanges = max_value - serise_rankChanges
            serise_rankChanges = serise_rankChanges / max_value
            
            serise_rankChanges.name = f'Rank_{propertyName}'

            return serise_rankChanges
        
        except Exception as e:
            return pd.Series()
        
    def get_ATRS150_exp_Ranks(self, Symbol):

        propertyName = 'ATRS150_Exp'

        try:
            rank_df = self.get_ATRS_Ranking_df()
            serise_rankChanges = rank_df.loc[Symbol]
            serise_rankChanges.name = f'Rank_{propertyName}'

            return serise_rankChanges
        
        except Exception as e:
            return pd.Series()
        
    def get_ATRS_Ranking_df(self):

        propertyName = 'ATRS150_Exp'

        try:
            if self.atrs_ranking_df.empty == True:
                csv_path = os.path.join(METADATA_FOLDER, f'{propertyName}_Ranking.csv')
                rank_df = pd.read_csv(csv_path)
                rank_df = rank_df.set_index('Symbol')
                self.atrs_ranking_df = rank_df
                return rank_df
            else:
                return self.atrs_ranking_df
        
        except Exception as e:
            return pd.DataFrame()
        
    def cook_Stock_GICS_df(self):
        """
        If you got the HTTP 404 error, always check  yf library version first.
        cmd: pip install --upgrade yahooquery

        sometimes there's invalid sector and industry data in the csv file because of the yf's worng database.
        in that case, just modify file in the excel editor.

        """
        all_list = self.get_local_stock_list()
        symbols = all_list['Symbol'].tolist()
        df_list = []
        requestQueue = []
        symbolsNum = len(symbols)
        errorTickers = []
        for i in range(0, symbolsNum):
            requestQueue.append(symbols[i])
            if len(requestQueue) >= 10:
                try:
                    tickers = Ticker(requestQueue)
                    profiles = tickers.get_modules("summaryProfile")
                    df = pd.DataFrame.from_dict(profiles).T
                    sector =  df['sector']
                    industry = df['industry']
                    df = pd.concat([sector, industry], axis=1)
                    df.index.name = 'Symbol'
                    df_list.append(df)
                    requestQueue.clear()
                    print(f"{i/symbolsNum*100:.2f}% Done")
                except Exception as e:
                    print(e)
                    print(requestQueue)
                    errorTickers.extend(requestQueue)
                    requestQueue.clear()
                    

        if len(requestQueue) > 0:
            try:
                tickers = Ticker(requestQueue)
                profiles = tickers.get_modules("summaryProfile")
                df = pd.DataFrame.from_dict(profiles).T
                sector =  df['sector']
                industry = df['industry']
                df = pd.concat([sector, industry], axis=1)
                df.index.name = 'Symbol'
                df_list.append(df)
                requestQueue.clear()
            except Exception as e:
                print(e)
                print(requestQueue)
                errorTickers.extend(requestQueue)
                requestQueue.clear()


        errorSymbolNum = len(errorTickers)
        symbols = errorTickers.copy()
        errorTickers.clear()
        for i in range(0, errorSymbolNum):
            requestQueue.append(symbols[i])
            if len(requestQueue) >= 1:
                try:
                    tickers = Ticker(requestQueue)
                    profiles = tickers.get_modules("summaryProfile")
                    df = pd.DataFrame.from_dict(profiles).T
                    sector =  df['sector']
                    industry = df['industry']
                    df = pd.concat([sector, industry], axis=1)
                    df.index.name = 'Symbol'
                    df_list.append(df)
                    requestQueue.clear()
                    print(f"{i/symbolsNum*100:.2f}% Done")
                except Exception as e:
                    print(e)
                    print(requestQueue)
                    errorTickers.extend(requestQueue)
                    requestQueue.clear()
                time.sleep(1)



        print('error tickers can\'t get the info')
        print(errorTickers)

        print("Complete!")

        result_df = pd.concat(df_list)
        # remove duplicated df
        result_df = result_df.loc[~result_df.index.duplicated()]
        
        result_df.index.name = 'Symbol'
        result_df.to_csv(os.path.join(METADATA_FOLDER, 'Stock_GICS.csv'), encoding='utf-8-sig')

        print('Error Tickers: ', errorTickers)


        return result_df
    

    def get_GICS_df(self):
        if self.stock_GICS_df.empty:
            # TODO: 예외처리 추가
            csv_path = os.path.join(METADATA_FOLDER, "Stock_GICS.csv")
            self.stock_GICS_df = pd.read_csv(csv_path)
            self.stock_GICS_df = self.stock_GICS_df.set_index('Symbol')
        
        return self.stock_GICS_df

   
    # method: industry or sector
    # N_day_before: rank will be calculated at the [:, -N_day_before].
    def get_industry_atrs150_ranks_mean(self, ATRS_Ranks_df, stock_GICS_df, N_day_before = 1, method : str = 'industry'):
        category = method

        tickers_by_category = stock_GICS_df.groupby(category)['Symbol'].apply(list).to_dict()

        # ex) -1, -5, -20, -60, -120, -240 (오늘, 일주전, 한달전, 3개월전, 6개월전, 1년전)
        last_ranks = ATRS_Ranks_df.iloc[:, -N_day_before]
        sectorTotalScoreMap = {}
        for category, tickers in tickers_by_category.items():
            scores = []
            # collect each sector's scores and dump lower 50%
            for ticker in tickers:          
                score = last_ranks.get(ticker) 
                if score is not None:
                    scores.append(score)

            scores = [score for score in scores if not np.isnan(score)]
            scores = sorted(scores)
            half_index = len(scores) // 2 # to dump half of scores
            half_dump_sector_scores = scores[:half_index] # dump half of scores. use 0 to halt_index (bigger numbers, bad ranks)
            length = len(half_dump_sector_scores)
            if length != 0:
                sectorTotalScoreMap[category] = sum(half_dump_sector_scores) / length

        # sort sector scores. the lower, the better
        sorted_sector_scores = dict(sorted(sectorTotalScoreMap.items(), key=lambda x: x[1]))
        return sorted_sector_scores
    

    def get_ranks_in_industries(self, ATRS_Ranks_df, stock_GICS_df):
        tickers_by_category = stock_GICS_df.groupby('industry')['Symbol'].apply(list).to_dict()

        # get last ranks
        last_ranks = ATRS_Ranks_df.iloc[:, -1]
        sorted_industries_ranks_dic = {}
        for category, tickers in tickers_by_category.items():
            # generate sorted ticker-rank dictionary
            industry_ranks = {ticker: last_ranks.get(ticker) for ticker in tickers if not pd.isna(last_ranks.get(ticker))}
            sorted_ranks = sorted(industry_ranks.items(), key=lambda x: x[1])
            sorted_industries_ranks_dic[category] = sorted_ranks

        return sorted_industries_ranks_dic

    # cook industry ranks according to the ATRS150_Exp ranks.
    def cook_long_term_industry_rank_scores(self):
        print('cook_long_term_industry_rank_scores')
        ATRS_Ranks_df = self.get_ATRS_Ranking_df()
        csv_path = os.path.join(METADATA_FOLDER, "Stock_GICS.csv")
        stock_GICS_df = pd.read_csv(csv_path)


        dn_list = []
        columnNames = []
        nDay = 365
        for i in range(1, nDay+1):
            dn = self.get_industry_atrs150_ranks_mean(ATRS_Ranks_df, stock_GICS_df, i)
            dn_list.append(dn)
            columnNames.append(f'{i}d ago')

        # reverse the list so that lastest day can be located at the last column in dataframe.
        dn_list.reverse()
        columnNames.reverse()

        d1 = dn_list[0]
        industryNames = list(d1.keys())
        industryNum = len(industryNames)
        industry_rank_history_dic = {}

        # generate industryName-list dictionary.
        for name in industryNames:
            industry_rank_history_dic[name] = []

        for i in range(1, nDay+1):
            for key, value in dn_list[i-1].items():
                try:
                    industry_rank_history_dic[key].append(value)
                except Exception as e:
                    print(e)
        rank_history_df = pd.DataFrame.from_dict(industry_rank_history_dic).transpose()
        rank_history_df = rank_history_df.rank(axis=0, ascending=True, method='dense')
        rank_history_df = (1 - rank_history_df.div(industryNum))*100
        rank_history_df = rank_history_df.round(2)
        rank_history_df.columns = columnNames
        rank_history_df.index.name = 'industry'


        # sort ranks by lastest score.
        last_col_name = rank_history_df.columns[-1]
        rank_history_df = rank_history_df.sort_values(by=last_col_name, ascending=False)


        save_path = os.path.join(METADATA_FOLDER, "long_term_industry_rank_scores.csv")
        rank_history_df.index.name = 'industry'
        rank_history_df.to_csv(save_path, encoding='utf-8-sig')
        print('long_term_industry_rank_scores.csv cooked!')

    def get_industry_atrs14_mean(self, stock_data_dic, stock_GICS_df, N_day_before = 1, method : str = 'industry'):
            category = method

            tickers_by_category = stock_GICS_df.groupby(category)['Symbol'].apply(list).to_dict()

            # ex) -1, -5, -20, -60, -120, -240 (오늘, 일주전, 한달전, 3개월전, 6개월전, 1년전)    
            sectorTotalScoreMap = {}
            for category, tickers in tickers_by_category.items():
                scores = []
                # collect each sector's atrs and dump lower 50%
                for ticker in tickers:     
                    df = stock_data_dic.get(ticker)           
                    if df is not None and (len(df) > N_day_before):
                        scores.append(df['ATRS_Exp'].iloc[-N_day_before])
                # ascending order sort
                scores = [score for score in scores if not np.isnan(score)]
                scores = sorted(scores)
                half_index = len(scores) // 2 # to dump half of scores get the index of middle
                half_dump_sector_scores = scores[half_index:] #  dump half of scores. Use from half_index to last (lower numbers)
                length = len(half_dump_sector_scores)
                if length != 0:
                    sectorTotalScoreMap[category] = sum(half_dump_sector_scores) / length

            # sort sector scores. The bigger atrs14, the better 
            sorted_sector_scores = dict(sorted(sectorTotalScoreMap.items(), key=lambda x: x[1], reverse=True))
            return sorted_sector_scores
    

    def get_percentage_AtoB(self, priceA : float, priceB : float):
        res = ((priceB - priceA)/priceA) * 100
        return res


    def get_DCR_normalized(self, inStockData: pd.DataFrame, n_day_before = -1):      
        # [DCR](%)

        d0_close = inStockData['Close'].iloc[n_day_before]
        d0_low = inStockData['Low'].iloc[n_day_before]
        d0_high = inStockData['High'].iloc[n_day_before]

        if d0_high - d0_low > 0:
            DCR = (d0_close - d0_low) / (d0_high - d0_low)
        else:
            DCR = 0.0

        return DCR

    # ------------------------------------------------
    # [ Public Helper ]: 단일 티커 날짜 탐색 및 로딩
    # ------------------------------------------------
    def find_days_by_return(self, ticker: str, pct_threshold: float = -3.0, days_num: int = 365*5) -> pd.DataFrame:
        """
        지정한 티커의 일일 등락률(%)이 pct_threshold 이하인 날짜를 반환합니다.
        예) pct_threshold = -3.0  => -3% 이상 하락한 날만 필터링

        반환 컬럼: ['Open','High','Low','Close','Volume','change_pct']
        """
        df = self.load_single_ticker_data(ticker, days_num)
        if df.empty or 'Close' not in df:
            return pd.DataFrame()

        df = df.sort_index()
        if len(df) < 2:
            return pd.DataFrame()

        prev_close = df['Close'].shift(1)
        df['change_pct'] = (df['Close'] / prev_close - 1.0) * 100.0

        res = df[df['change_pct'] <= pct_threshold].copy()
        if res.empty:
            return pd.DataFrame()

        return res[['Open','High','Low','Close','Volume','change_pct']]

    def find_days_drop_and_ema_violation(
        self,
        ticker: str,
        pct_threshold: float = -3.0,
        atr_multiple: float = 2.0,
        days_num: int = 365*5
    ) -> dict:
        """
        단일 티커에 대해 "EMA 리젝션(거부)" 조건과 함께 두 가지 하락 조건을 탐지합니다.

        정의 및 계산식:
        - EMA 리젝션: 시가가 10EMA, 21EMA 모두 위에서 시작하고, 종가가 10EMA, 21EMA 모두 아래에서 마감
          (Open > EMA10, Open > EMA21) AND (Close < EMA10, Close < EMA21)
        - 퍼센트 하락: change_pct = (Close / PrevClose - 1) * 100 <= pct_threshold
        - ATR 배수 하락: ATR_Drop = (Open - Close) / ATR(10) >= abs(atr_multiple)
          (여기서 ATR은 True Range의 10일 이동평균)

        처리 절차:
        1) 최근 days_num 기간의 단일 티커 데이터를 로드하고 날짜 기준 정렬
        2) EMA10, EMA21, ATR(10) 계산
        3) 전일 종가(PrevClose)로부터 change_pct 및 ATR_Drop 계산
        4) EMA 리젝션 조건과 각 하락 조건(퍼센트/ATR)을 결합하여 필터링

        반환값(dict):
        - 'drop_pct': EMA 리젝션이면서 퍼센트 하락 조건을 만족
        - 'drop_atr': EMA 리젝션이면서 ATR 배수 하락 조건을 만족
        - 'intersection': 위 두 조건을 모두 만족하는 교집합

        반환 DataFrame 공통 컬럼:
        ['Open','High','Low','Close','Volume','EMA10','EMA21','ATR','change_pct','ATR_Drop']
        """

        df = self.load_single_ticker_data(ticker, days_num)
        if df.empty or 'Close' not in df:
            return {
                'drop_pct': pd.DataFrame(),
                'drop_atr': pd.DataFrame(),
                'intersection': pd.DataFrame()
            }

        df = df.sort_index()

        # Indicators
        prev_close = df['Close'].shift(1)
        df['EMA10'] = jdi.ema(df['Close'], span=10)
        df['EMA21'] = jdi.ema(df['Close'], span=21)
        df['ATR']   = jdi.atr(df, 10)

        # Changes
        df['change_pct'] = (df['Close'] / prev_close - 1.0) * 100.0
        
        # ATR 하락폭(양수): 시가 대비 종가 하락폭을 ATR 배수로 환산
        df['ATR_Drop'] = (df['Open'] - df['Close']) / df['ATR']

        # EMA rejection: Open above both EMAs while Close ends below both
        ema_reject = (
            (df['Open'] > df['EMA10']) & (df['Open'] > df['EMA21']) &
            (df['Close'] < df['EMA10']) & (df['Close'] < df['EMA21'])
        )

        cond_pct = (df['change_pct'] <= pct_threshold) & ema_reject
        cond_atr = (df['ATR_Drop'] >= abs(atr_multiple)) & ema_reject
        cond_both = cond_pct & cond_atr

        cols = ['Open','High','Low','Close','Volume','EMA10','EMA21','ATR','change_pct','ATR_Drop']
        return {
            'drop_pct': df.loc[cond_pct, cols].copy(),
            'drop_atr': df.loc[cond_atr, cols].copy(),
            'intersection': df.loc[cond_both, cols].copy(),
        }

    def load_single_ticker_data(self, ticker: str, days_num: int = 365*5) -> pd.DataFrame:
        """
        단일 티커 데이터를 로드합니다.
         - 1순위: 로컬 CSV(StockData/{ticker}.csv)
         - 2순위: FinanceDataReader에서 직접 수집(ETF/지수 티커 포함, 예: QQQ, SPY)

        반환: Date 인덱스 정렬된 DataFrame (가능 시 ['Open','High','Low','Close','Volume'] 포함)
        """
        try:
            today_date = dt.date.today()
            start_date = today_date - dt.timedelta(days=days_num)

            # 로컬 CSV 우선
            local_df = load_csv_with_date_index(
                ticker, data_dir=DATA_FOLDER, start_date=start_date, end_date=today_date
            )
            if not local_df.empty:
                return local_df.sort_index()

            # 로컬에 없으면 FDR에서 수집 (ETF/지수 포함)
            try:
                fetched = fdr.DataReader(ticker, start_date)
            except Exception as e:
                logging.info(f"[load_single_ticker_data] FDR fetch failed for {ticker}: {e}")
                return pd.DataFrame()

            if fetched is None or fetched.empty:
                return pd.DataFrame()

            # Date 인덱스 정렬 및 표준 컬럼 보존
            fetched = fetched.sort_index()
            if 'Volume' not in fetched.columns:
                fetched['Volume'] = 0

            needed = ['Open','High','Low','Close','Volume']
            # Close는 반드시 필요. 없으면 빈 DF 반환
            if 'Close' not in fetched.columns:
                return pd.DataFrame()
            for c in needed:
                if c not in fetched.columns:
                    fetched[c] = 0

            return fetched[needed]

        except Exception as e:
            logging.info(f"[load_single_ticker_data] Unexpected error for {ticker}: {e}")
            return pd.DataFrame()
        

    # cook industry ranks according to the atrs14_exp.
    def cook_short_term_industry_rank_scores(self):
        print('cook_short_term_industry_rank_scores')

        all_list = self.get_local_stock_list()
        stock_datas_from_csv = self.get_stock_datas_from_csv(all_list, 365, True)

        csv_path = os.path.join(METADATA_FOLDER, "Stock_GICS.csv")
        stock_GICS_df = pd.read_csv(csv_path)


        dn_list = []
        columnNames = []
        nDay = 30
        for i in range(1, nDay+1):
            dn = self.get_industry_atrs14_mean(stock_datas_from_csv, stock_GICS_df, i, 'industry')
            dn_list.append(dn)
            columnNames.append(f'{i}d ago')


        # reverse the list so that lastest day can be located at the last column in dataframe.
        dn_list.reverse()
        columnNames.reverse()

        d1 = dn_list[0]
        industryNames = list(d1.keys())
        industryNum = len(industryNames)
        industry_atrs14_rank_history_dic = {}

        # generate industryName-list dictionary.
        for name in industryNames:
            industry_atrs14_rank_history_dic[name] = []

        for i in range(1, nDay+1):
            for key, value in dn_list[i-1].items():
                try:
                    industry_atrs14_rank_history_dic[key].append(value)
                except Exception as e:
                    print(e)

        rank_history_df = pd.DataFrame.from_dict(industry_atrs14_rank_history_dic).transpose()
        rank_history_df = rank_history_df.rank(axis=0, ascending=False, method='dense')
        rank_history_df = (1 - rank_history_df.div(industryNum))*100
        rank_history_df = rank_history_df.round(2)
        rank_history_df.columns = columnNames
        rank_history_df.index.name = 'industry'

        # sort ranks by lastest score.
        last_col_name = rank_history_df.columns[-1]
        rank_history_df = rank_history_df.sort_values(by=last_col_name, ascending=False)

        save_path = os.path.join(METADATA_FOLDER, "short_term_industry_rank_scores.csv")
        rank_history_df.to_csv(save_path, encoding='utf-8-sig')
        print('short_term_industry_rank_scores.csv cooked!')

    def check_NR_with_TrueRange(self, inStockData : pd.DataFrame , maxDepth = 20):
        # Calcaute NR(x) using True Range.
        last_tr = inStockData['TR'].iloc[-1]
        trueRange_NR_x = 0
        for i in range(2, maxDepth):
            tr_rangeN = inStockData['TR'][-i:]
            min_value = tr_rangeN.min()
            if min_value == last_tr:
                trueRange_NR_x = i
        return trueRange_NR_x
    
    def check_insideBar(self, inStockData: pd.DataFrame):
        d2_ago_high, d2_ago_low = inStockData['High'].iloc[-2], inStockData['Low'].iloc[-2]
        d1_ago_high, d1_ago_low = inStockData['High'].iloc[-1], inStockData['Low'].iloc[-1]
        bIsInsideBar = False
        bDoubleInsideBar = False
        if d2_ago_high > d1_ago_high and d2_ago_low < d1_ago_low:
            bIsInsideBar = True

        if bIsInsideBar:
            d3_ago_high, d3_ago_low = inStockData['High'].iloc[-3], inStockData['Low'].iloc[-3]
            if d3_ago_high > d2_ago_high and d3_ago_low < d2_ago_low:
                bDoubleInsideBar = True



        
        return bIsInsideBar, bDoubleInsideBar
    
    def get_moving_average_data(self, inStockData: pd.DataFrame, Num):
        return inStockData['Close'].rolling(window=Num).mean()

    # return boolean tuple (bIsConverging, bIsPower3)
    def check_ma_converging(self, inStockData: pd.DataFrame):

        bIsConverging = False
        bIsPower3 = False
        bIsPower2 = False

        ma10_datas = inStockData['Close'].rolling(window=10).mean()
        ma20_datas = inStockData['Close'].rolling(window=20).mean()
        ma50_datas = inStockData['Close'].rolling(window=50).mean()

        ma10 = ma10_datas.iloc[-1]
        ma20 = ma20_datas.iloc[-1]
        ma50 = ma50_datas.iloc[-1]

        dist_10_20 =  abs((ma10 - ma20)/ma20) * 100
        dist_10_50 =  abs((ma10 - ma50)/ma50) * 100
        dist_20_50 =  abs((ma50 - ma20)/ma20) * 100

        if dist_10_20 < 1.5 and dist_10_50 < 1.5 and dist_20_50 < 1.5:
            bIsConverging = True

        if bIsConverging:
            low = inStockData['Low'].iloc[-1]
            close = inStockData['Close'].iloc[-1]

            ma_list = [ma10, ma20]
            ma_min = min(ma_list)
            ma_max = max(ma_list)

            if low < ma_min and close > ma_max:
                bIsPower2 = True
           
            ma_list = [ma10, ma20, ma50]
            ma_min = min(ma_list)
            ma_max = max(ma_list)

            if low < ma_min and close > ma_max:
                bIsPower3 = True

        return (bIsConverging, bIsPower3, bIsPower2)

    
    # check the low or close was closed to the moving average
    def check_near_ma(self, inStockData: pd.DataFrame, MA_Num = 10, max_error_pct = 1.5, bUseEMA = False):

        if bUseEMA:
            ma_datas = inStockData['Close'].ewm(span=MA_Num, adjust=False).mean()
        else:
            ma_datas = inStockData['Close'].rolling(window=MA_Num).mean()

        low = inStockData['Low'].iloc[-1]
        close = inStockData['Close'].iloc[-1]
        ma = ma_datas.iloc[-1]

        

        ma_dist_from_close = abs(self.get_percentage_AtoB(close, ma))
        ma_dist_from_low = abs(self.get_percentage_AtoB(low, ma))

        return ma_dist_from_close < max_error_pct or ma_dist_from_low < max_error_pct
   

    def check_supported_by_ma(self, inStockData: pd.DataFrame, MA_Num = 10, max_error_pct = 1.5, bUseEMA = False):
        # print('low > ma and low and ma dist < max_dist_pct')
        # print('low < ma and close > ma')

        if bUseEMA:
            ma_datas = inStockData['Close'].ewm(span=MA_Num, adjust=False).mean()
        else:
            ma_datas = inStockData['Close'].rolling(window=MA_Num).mean()

        
        low = inStockData['Low'].iloc[-1]
        close = inStockData['Close'].iloc[-1]
        ma = ma_datas.iloc[-1]

        dist_from_close = abs(self.get_percentage_AtoB(close, ma))
        dist_from_low = abs(self.get_percentage_AtoB(low, ma))


        if low > ma and dist_from_low < max_error_pct:
            return True
        if low < ma and close > ma:
            return True

        return False
    

    def check_ma_touch(self, inStockData: pd.DataFrame, MA_Num = 10, bUseEMA = False, n_day_before = -1):
        """
        check if the ma price is in the day's range.
        """
        if bUseEMA:
            ma_datas = inStockData['Close'].ewm(span=MA_Num, adjust=False).mean()
        else:
            ma_datas = inStockData['Close'].rolling(window=MA_Num).mean()
     
        low = inStockData['Low'].iloc[n_day_before]
        high = inStockData['High'].iloc[n_day_before]
        ma = ma_datas.iloc[n_day_before]

        if ma <= high and ma >= low:
            return True
        
        return False


    def check_undercut_price(self, inStockData: pd.DataFrame, inPrice : float, n_day_before = -1):
        """
        check if the day's low price undercut the {price}.
        """
        low = inStockData['Low'].iloc[n_day_before]
        if low < inPrice:
            return True
        
        return False


    def check_wickplay(self, inStockData: pd.DataFrame):
        bWickPlay = False
        d2_ago_open, d2_ago_high, d2_ago_low, d2_ago_close = inStockData['Open'].iloc[-2], inStockData['High'].iloc[-2], inStockData['Low'].iloc[-2], inStockData['Close'].iloc[-2]
        d1_ago_open, d1_ago_high, d1_ago_low, d1_ago_close = inStockData['Open'].iloc[-1], inStockData['High'].iloc[-1], inStockData['Low'].iloc[-1], inStockData['Close'].iloc[-1]

        # bullish candle
        if d2_ago_open <= d2_ago_close:
            if d1_ago_high <= d2_ago_high and d1_ago_low >= d2_ago_close:
                bWickPlay = True
        # bearish candle
        else:
            if d1_ago_high <= d2_ago_high and d1_ago_low >= d2_ago_open:
                bWickPlay = True

        return bWickPlay
    

    # open equal low
    def check_OEL(self, inStockData: pd.DataFrame, n_day_before = -1):
        d1_ago_open, d1_ago_low = inStockData['Open'].iloc[n_day_before], inStockData['Low'].iloc[n_day_before]
        bOEL = d1_ago_open == d1_ago_low
        return bOEL
    
    def check_OEH(self, inStockData: pd.DataFrame, n_day_before = -1):
        d1_ago_open, d1_ago_high = inStockData['Open'].iloc[n_day_before], inStockData['High'].iloc[n_day_before]
        bOEH = d1_ago_open == d1_ago_high
        return bOEH
    



    
    


    # --------------- C/V Check list --------------
    def check_lower_lows_3(self, inStockData: pd.DataFrame, days=15):
        ll_cnt = 0
        ticker = inStockData['Symbol'].iloc[-1]
        lows = inStockData['Low'].iloc[-days:].tolist()
        closes = inStockData['Close'].iloc[-days:].tolist()

        for i in range(0, days-1):
            # lower low
            prev_low = lows[i]
            today_close = closes[i+1]
            if prev_low > today_close:
                ll_cnt += 1
            else:
                ll_cnt = 0

            if ll_cnt >= 3:
                return True
        return False
    

    def check_higher_highs_3(self, inStockData: pd.DataFrame, days=15):
        hh_cnt = 0
        ticker = inStockData['Symbol'].iloc[-1]
        highs = inStockData['High'].iloc[-days:].tolist()
        closes = inStockData['Close'].iloc[-days:].tolist()

        for i in range(0, days-1):
            prev_high = highs[i]
            today_close = closes[i+1]
            # higher high
            if prev_high < today_close:
                hh_cnt += 1
            else:
                hh_cnt = 0
            if hh_cnt >= 3:
                return True

        return False
    

    def _check_below_N_ma_closed_cnt(self, inStockData: pd.DataFrame, MA_Num, days):

        ticker = inStockData['Symbol'].iloc[-1]
        ma_datas = inStockData['Close'].rolling(window=MA_Num).mean().iloc[-days:].tolist()
        opens = inStockData['Open'].iloc[-days:].tolist()
        closes = inStockData['Close'].iloc[-days:].tolist()

        ma_below_closed_cnt = 0
        for i in range(0, days):
            ma = ma_datas[i]
            open = opens[i]
            close = closes[i]
            if open >= ma and close < ma:
                ma_below_closed_cnt += 1
            

        return ma_below_closed_cnt


    def check_below_20ma_closed_cnt(self, inStockData: pd.DataFrame, days=15):
        return self._check_below_N_ma_closed_cnt(inStockData, 20, days)

    def check_below_50ma_closed_cnt(self, inStockData: pd.DataFrame, days=15):
        return self._check_below_N_ma_closed_cnt(inStockData, 50, days)
    
    def check_up_more_than_adr_cnt(self, inStockData: pd.DataFrame, days=15):
        ticker = inStockData['Symbol'].iloc[-1]
        closes = inStockData['Close'].iloc[-days:].tolist()
        ADRs = inStockData['ADR'].iloc[-days:].tolist()

        cnt = 0
        for i in range(0, days-1):
            prev_close = closes[i]
            close = closes[i+1]
            dist_percentage = self.get_percentage_AtoB(prev_close, close)
            if dist_percentage > 0 and abs(dist_percentage) >= ADRs[i]:
                cnt += 1
            
        return cnt
    
    def check_down_more_than_adr_cnt(self, inStockData: pd.DataFrame, days=15):
        ticker = inStockData['Symbol'].iloc[-1]
        closes = inStockData['Close'].iloc[-days:].tolist()
        ADRs = inStockData['ADR'].iloc[-days:].tolist()

        cnt = 0
        for i in range(0, days-1):
            prev_close = closes[i]
            close = closes[i+1]
            dist_percentage = self.get_percentage_AtoB(prev_close, close)
            if dist_percentage < 0 and abs(dist_percentage) >= ADRs[i]:
                cnt += 1
            
        return cnt
    
    def check_bullish_bearish_candle_count_in_n_days(self, inStockData: pd.DataFrame, days=15):
        opens = inStockData['Open'].iloc[-days:].tolist()
        closes = inStockData['Close'].iloc[-days:].tolist()

        bullishCandleCnt = 0
        bearlishCandleCnt = 0

        for i in range(0, days):
            open = opens[i]
            close = closes[i]

            if open < close:
                bullishCandleCnt += 1
            else:
                bearlishCandleCnt += 1


        return bullishCandleCnt, bearlishCandleCnt
    

    def check_20ma_dist_more_than_20ptg_cnt(self, inStockData: pd.DataFrame, days=15):
        closes = inStockData['Close'].iloc[-days:].tolist()
        ma20_prices = inStockData['Close'].rolling(window=20).mean().iloc[-days:].tolist()

        big_dist_cnt = 0
        for i in range(0, days):   
            close = closes[i]
            ma20 = ma20_prices[i]

            if close > ma20:
                dist_from_20ma =  abs((ma20 - close)/close) * 100
                if dist_from_20ma > 20:
                    big_dist_cnt += 1

        return big_dist_cnt
    

    def check_close_equal_high_or_low_cnt(self, inStockData: pd.DataFrame, days=15):
        closes = inStockData['Close'].iloc[-days:].tolist()
        highs = inStockData['High'].iloc[-days:].tolist()
        lows = inStockData['Low'].iloc[-days:].tolist()
  
        close_equal_high_cnt = 0
        close_equal_low_cnt = 0


        for i in range(0, days):
            close = closes[i]
            high = highs[i]
            low = lows[i]

            if low != high:
                if close == high:
                    close_equal_high_cnt += 1
                if close == low:
                    close_equal_low_cnt += 1


        return close_equal_high_cnt, close_equal_low_cnt
    

    # gap + OEL
    def check_GOEL(self, inStockData: pd.DataFrame, n_day_before = -1):
        
        d1_ago_close = inStockData['Close'].iloc[n_day_before - 1]
        d1_ago_high = inStockData['High'].iloc[n_day_before - 1]
        d0_open =  inStockData['Open'].iloc[n_day_before]

        # Gap Start
        if d1_ago_high < d0_open:
            # more than 3% up
            if self.get_percentage_AtoB(d1_ago_close, d0_open) >= 3.0:
                # and OEL
                if self.check_OEL(inStockData, n_day_before):
                    return True
    
        return False

    def check_failed_downside_wick_BO(self, inStockData: pd.DataFrame):
        d2_ago_open, d2_ago_high, d2_ago_low, d2_ago_close = inStockData['Open'].iloc[-2], inStockData['High'].iloc[-2], inStockData['Low'].iloc[-2], inStockData['Close'].iloc[-2]
        d1_ago_open, d1_ago_high, d1_ago_low, d1_ago_close = inStockData['Open'].iloc[-1], inStockData['High'].iloc[-1], inStockData['Low'].iloc[-1], inStockData['Close'].iloc[-1]

        # bullish candle
        if d2_ago_open <= d2_ago_close:
            # 전일 심지 범위 안에서 출발
            if d1_ago_open < d2_ago_open and d1_ago_open > d2_ago_low:
                # 장중에 downside wick BO 발생
                if d1_ago_low < d2_ago_low:
                    # But 이말올 하여 종가는 전날 심지(open) 위에서 마감
                    if d1_ago_close > d2_ago_open:
                        return True
                    
        # bearish candle
        else:
            # 전일 심지 아래에서 출발
            if d1_ago_open < d2_ago_close and d1_ago_open > d2_ago_low:
                # 장중에 downside wick BO 발생
                if d1_ago_low < d2_ago_low:
                    # But 이말올하여 종가는 전날 심지(종가) 위에서 마감
                    if d1_ago_close > d2_ago_close:
                        return True
                

        return False
    


    def check_oops_up_reversal(self, inStockData: pd.DataFrame):
        low_d1_ago = inStockData['Low'].iloc[-2]
        open = inStockData['Open'].iloc[-1]
        close = inStockData['Close'].iloc[-1]

        if open < low_d1_ago and close > low_d1_ago:
            return True
        
        return False


    def check_oops_up_reversal_cnt(self, inStockData: pd.DataFrame, days=15):
        opens = inStockData['Open'].iloc[-days:].tolist()
        closes = inStockData['Close'].iloc[-days:].tolist()
        lows = inStockData['Low'].iloc[-days:].tolist()

        oops_up_cnt = 0

        for i in range(0, days-1):
            prev_low = lows[i]
            open = opens[i+1]
            close = closes[i+1]

            if open < prev_low and close > prev_low:
                oops_up_cnt += 1

        return oops_up_cnt


    def check_OEL_cnt(self, inStockData: pd.DataFrame, days=15):
        ticker = inStockData['Symbol'].iloc[-1]
        opens = inStockData['Open'].iloc[-days:].tolist()
        lows = inStockData['Low'].iloc[-days:].tolist()

        condition_cnt = 0
        for i in range(0, days):
            open = opens[i]
            low = lows[i]
            if open == low:
                condition_cnt += 1
    
        return condition_cnt
    

    def check_OEH_cnt(self, inStockData: pd.DataFrame, days=15):
        ticker = inStockData['Symbol'].iloc[-1]
        opens = inStockData['Open'].iloc[-days:].tolist()
        highs = inStockData['High'].iloc[-days:].tolist()

        condition_cnt = 0
        for i in range(0, days):
            open = opens[i]
            high = highs[i]
            if open == high:
                condition_cnt += 1

        return condition_cnt
    

    # 전일 고가보다 위로 1 ADR% 만큼 주가가 상승했으나 전일 고가 아래에서 마감
    def check_squat_cnt(self, inStockData: pd.DataFrame, days=15):
        ticker = inStockData['Symbol'].iloc[-1]
        highs = inStockData['High'].iloc[-days:].tolist()
        closes = inStockData['Close'].iloc[-days:].tolist()
        ADRs = inStockData['ADR'].iloc[-days:].tolist()

        squat_cnt = 0
        for i in range(0, days-1):
            prevHigh = highs[i]
            high = highs[i+1]
            close = closes[i+1]
            adr = ADRs[i]

            if high > prevHigh:
                dist_pct = self.get_percentage_AtoB(prevHigh, high)
                if dist_pct >= adr:
                    if close < prevHigh:
                        squat_cnt += 1

        return squat_cnt
    
    # 스쿼트 발생 이후 3일 안에 회복 (스쿼트가 먼저 발생해야 한다.)
    def check_squat_recovery_cnt(self, inStockData: pd.DataFrame, days=15):
        ticker = inStockData['Symbol'].iloc[-1]
        highs = inStockData['High'].iloc[-days:].tolist()
        closes = inStockData['Close'].iloc[-days:].tolist()
        ADRs = inStockData['ADR'].iloc[-days:].tolist()

        recovery_success_cnt = 0
        for i in range(0, days-1):
            prevHigh = highs[i]
            high = highs[i+1]
            close = closes[i+1]
            adr = ADRs[i]

            if high > prevHigh:
                dist_pct = self.get_percentage_AtoB(prevHigh, high)
                if dist_pct >= adr:
                    # 스쿼트 발생!
                    if close < prevHigh:
                        # 스쿼트 발생후 3일 이내 회복 확인
                        # 1d after squat
                        if i+2 < days:
                            close_after_squat_1d = closes[i+2]
                            if high < close_after_squat_1d:
                                recovery_success_cnt += 1
                                continue

                        # 2d after squat
                        if i+3 < days:
                            close_after_squat_2d = closes[i+3]
                            if high < close_after_squat_2d:
                                recovery_success_cnt += 1
                                continue

                        # 3d after squat
                        if i+4 < days:
                            close_after_squat_3d = closes[i+4]
                            if high < close_after_squat_3d:
                                recovery_success_cnt += 1
                                continue
                    

        return recovery_success_cnt
    

    # rs_check_range: Check if RS is maximum within N days.
    # days: Check if the day with the maximum RS occurred among n days.
    def check_rs_N_day_new_high_in_n_days(self, inStockData: pd.DataFrame, atrs_ranking_df : pd.DataFrame, rs_check_range = 50, days = 15):
        ticker = inStockData['Symbol'].iloc[-1]
        rs_of_ticker = atrs_ranking_df.loc[ticker]
        rs_ranks_in_n_days = rs_of_ticker.iloc[-days:].tolist()

        for i in range(0, days):
            # [-days ~ 0]
            n_day_before_index = i - days
            range_day_before_from_index = n_day_before_index - rs_check_range
            today_rs_rank = rs_of_ticker.iloc[n_day_before_index]
            # get last 'rs_check_range' days RS ranks
            last_rs_ranks_in_range = rs_of_ticker.iloc[range_day_before_from_index : n_day_before_index]
            # get top rs rank in last 'rs_check_range' days.
            top_rank_in_last_range_days = last_rs_ranks_in_range.min()

            # rs rank new high.
            if today_rs_rank < top_rank_in_last_range_days:
                # print(ticker)
                # print(n_day_before_index, ' days before rs rank: ', today_rs_rank)
                # print('last ', rs_check_range, 'days top rs rank from ', n_day_before_index, 'days: ', top_rank_in_last_range_days)
                # print(today_rs_rank, ' < ', top_rank_in_last_range_days)
                return True

        return False
    
    def check_rs_N_day_new_low_in_n_days(self, inStockData: pd.DataFrame, atrs_ranking_df : pd.DataFrame, rs_check_range = 50, days = 15):
        ticker = inStockData['Symbol'].iloc[-1]
        rs_of_ticker = atrs_ranking_df.loc[ticker]
        rs_ranks_in_n_days = rs_of_ticker.iloc[-days:].tolist()

        for i in range(0, days):
            # [-days ~ 0]
            n_day_before_index = i - days
            range_day_before_from_index = n_day_before_index - rs_check_range
            today_rs_rank = rs_of_ticker.iloc[n_day_before_index]
            # get last 'rs_check_range' days RS ranks
            last_rs_ranks_in_range = rs_of_ticker.iloc[range_day_before_from_index : n_day_before_index]
            # get top rs rank in last 'rs_check_range' days.
            lowest_rank_in_last_range_days = last_rs_ranks_in_range.max()

            # rs rank new low.
            if today_rs_rank > lowest_rank_in_last_range_days:
                # print(ticker)
                # print(n_day_before_index, ' days before rs rank: ', today_rs_rank)
                # print('last ', rs_check_range, 'days lowest rs rank from ', n_day_before_index, 'days: ', lowest_rank_in_last_range_days)
                # print(today_rs_rank, ' > ', lowest_rank_in_last_range_days)
                return True

        return False
    



    def check_pocket_pivot(self, inStockData: pd.DataFrame):
        # Check Pocket pivot
        bIsPocketPivot = False

        # only bullish day
        bBullishDay = inStockData['Close'].iloc[-1] > inStockData['Open'].iloc[-1]
        if bBullishDay:
            # get the last 10 days from the last day. considering shift operaion below.
            recent_10_days_volumes = inStockData[-12:-1]
            last_day_volume = inStockData['Volume'].iloc[-1]

            # select price down days's volume
            price_drop_condition = recent_10_days_volumes['Close'] < recent_10_days_volumes['Close'].shift(1)
            volume_on_price_drop_days = recent_10_days_volumes.loc[price_drop_condition, 'Volume']

            # today's volume is bigger than last 10 down day's volume.
            if volume_on_price_drop_days.max() < last_day_volume:
                ma10 = self.get_moving_average_data(inStockData, 10).iloc[-1]
                low = inStockData['Low'].iloc[-1]
                ma10_to_low = self.get_percentage_AtoB(ma10, low)

                # stock price shouldn't above the ma10 more than 0.1%
                if ma10_to_low <= 0.1:
                    bIsPocketPivot = True

        return bIsPocketPivot


    def check_pocket_pivot_cnt(self, inStockData: pd.DataFrame, days=15):
        ticker = inStockData['Symbol'].iloc[-1]
        #dates = inStockData.iloc[-days:].index.tolist()
        closes = inStockData['Close'].iloc[-days:].tolist()
        opens = inStockData['Open'].iloc[-days:].tolist()
        volumes = inStockData['Volume'].iloc[-days:].tolist()
        lows = inStockData['Low'].iloc[-days:].tolist()
        ma10 = self.get_moving_average_data(inStockData, 10).iloc[-days:].tolist()
        pocket_pivot_cnt = 0

        for i in range(0, days):
            #today_date = dates[i]
            today_close = closes[i]
            today_open = opens[i]

            today_volume = volumes[i]
            today_low = lows[i]
            today_ma10 = ma10[i]

            bBullishDay = today_close > today_open
            # bullish day
            if bBullishDay:
                ma10_to_low_pcg = self.get_percentage_AtoB(today_ma10, today_low)
                # stock price shouldn't above the ma10 more than 0.1%
                if ma10_to_low_pcg <= 0.1:
                    n_day_before_index = i - days
                    last_10_days_volumes_range_index = n_day_before_index - 10
                    # [days -10, days]
                    last_10_days_volumes = inStockData[last_10_days_volumes_range_index : n_day_before_index]

                    # select price down day's volume
                    price_drop_condition = last_10_days_volumes['Close'] < last_10_days_volumes['Close'].shift(1)
                    volumes_of_price_drop_days = last_10_days_volumes.loc[price_drop_condition, 'Volume']

                    # today's day volume is bigger than last 10 down day's volume.
                    if volumes_of_price_drop_days.max() < today_volume:
                        pocket_pivot_cnt += 1

        return pocket_pivot_cnt

    # 현 주가가 150MA 및 200MA 위에 있는가?
    # 주가가 50일 MA위에 있는가?
    # 200MA, 150MA 기울기가 0보다 큰가?
    # 50MA가 150MA, 200MA 위로 상승하였는가?
    # 거래량이 거래하기 충분한가?

    # basically It's like a MMT. But ease the MA alignment condition.
    # + exclude low volume stocks
    def check_stage2(self, inStockData: pd.DataFrame, bOnly200MACheck = False):
            close = inStockData['Close'].iloc[-1]
            ma150 = inStockData['150MA'].iloc[-1]
            ma200 = inStockData['200MA'].iloc[-1]
            bIsUpperMA_150_200 = close > ma150 and close > ma200

            # early rejection for optimization
            if bOnly200MACheck == False:
                if bIsUpperMA_150_200 == False:
                    return False
            
            ma150_slope = inStockData['MA150_Slope'].iloc[-1]
            ma200_slope = inStockData['MA200_Slope'].iloc[-1]
            ma50 = inStockData['50MA'].iloc[-1]
            last_volume = inStockData['Volume'].iloc[-1]
            volume_ma50 = inStockData['Volume'].rolling(window=50).mean().iloc[-1]
            # (거래량평균 20만이상 + 10불이상 or 하루거래량 100억 이상) AND 마지막 거래량 10만주 이상
            bIsVolumeEnough = (volume_ma50 >= 200000 and close >= 10 ) or volume_ma50*close > 10000000
            bIsVolumeEnough = bIsVolumeEnough and last_volume >= 100000

            if bIsVolumeEnough == False:
                return False

            if bOnly200MACheck:
                bIsUpperMA = close > ma200

                filterMatchNum = 0
                if bIsUpperMA:
                    filterMatchNum = filterMatchNum + 1

                return filterMatchNum >= 1

            else:
                bIsUpperMA = close > bIsUpperMA_150_200 and close > ma50
                b_150ma_upper_than_200ma = ma150 > ma200
                b_50ma_biggerThan_150ma_200ma = ma50 > ma150 and ma50 > ma200
                bMA_Slope_Plus = ma150_slope > 0 and ma200_slope > 0

                filterMatchNum = 0

                if bIsUpperMA:
                    filterMatchNum = filterMatchNum + 1
                if b_150ma_upper_than_200ma or True:
                    filterMatchNum = filterMatchNum + 1
                if bMA_Slope_Plus:
                    filterMatchNum = filterMatchNum + 1
                if b_50ma_biggerThan_150ma_200ma:
                    filterMatchNum = filterMatchNum + 1

                return filterMatchNum >= 4



    def cook_top10_in_industries(self):
        ATRS_Ranks_df = self.get_ATRS_Ranking_df()
        csv_path = os.path.join(METADATA_FOLDER, "Stock_GICS.csv")
        stock_GICS_df = pd.read_csv(csv_path)
        ranks_in_industries = self.get_ranks_in_industries(ATRS_Ranks_df, stock_GICS_df)
        daysNum = 365
        stock_list = self.get_local_stock_list()
        stock_datas_from_csv = self.get_stock_datas_from_csv(stock_list, daysNum, False)

        top10_in_industries = {}
        for industry, datas in ranks_in_industries.items():
            top10_in_industry = []
            for data in datas:
                ticker, rank = data
                stock_data = stock_datas_from_csv.get(ticker, pd.DataFrame())
                if not stock_data.empty:
                    # just 200 MA check or MMT criteria
                    # bIsStage2 = self.check_stage2(stock_data, True)
                    # if bIsStage2:
                    top10_in_industry.append(data)

                    if len(top10_in_industry) == 10:
                        break

            top10_in_industries[industry] = top10_in_industry

        save_to_json(top10_in_industries, 'top10_in_industries')


    def get_long_term_industry_rank_scores_df(self):
        if self.long_term_industry_rank_df.empty:
            csv_path = os.path.join(METADATA_FOLDER, "long_term_industry_rank_scores.csv")
            self.long_term_industry_rank_df = pd.read_csv(csv_path)
            self.long_term_industry_rank_df = self.long_term_industry_rank_df.set_index('industry')

        return self.long_term_industry_rank_df
    
    def get_long_term_industry_rank_scores(self, industryName):
        df = self.get_long_term_industry_rank_scores_df()
        try:
            result = df.loc[industryName]
        except Exception as e:
            result = pd.Series()
            print(e)
            print(f"{industryName} does not exist in the DataFrame. Returning pd.Series().")

        return result

    def get_short_term_industry_rank_scores_df(self):
        if self.short_term_industry_rank_df.empty:
            csv_path = os.path.join(METADATA_FOLDER, "short_term_industry_rank_scores.csv")
            self.short_term_industry_rank_df = pd.read_csv(csv_path)
            self.short_term_industry_rank_df = self.short_term_industry_rank_df.set_index('industry')

        return self.short_term_industry_rank_df


        
        
    def get_top10_in_industries(self):
        dic = load_from_json('top10_in_industries')
        return dic
    
    def cook_stock_info_from_tickers(self, inTickers : list, fileName : str, bUseDataCache = True):
        cook_start_time = time.time()

        stock_list = self.get_local_stock_list()
        daysNum = 365
        stock_datas_from_csv = self.get_stock_datas_from_csv(stock_list, daysNum, bUseDataCache)
        atrs_ranking_df = self.get_ATRS_Ranking_df()

        nyse_list = self.data_getter.get_fdr_stock_list('NYSE')
        nyse_list = nyse_list['Symbol'].tolist()

        nasdaq_list = self.data_getter.get_fdr_stock_list('NASDAQ')
        nasdaq_list = nasdaq_list['Symbol'].tolist()

        stock_info_dic = {}

        for ticker in inTickers:
            gisc_df = self.get_GICS_df()
            market = ''

            if ticker in nyse_list:
                market = 'NYSE'
            if ticker in nasdaq_list:
                market = 'NASDAQ'

            if market == '':
                print('can not find ticker {0} in any nyse or nasdaq market.', ticker)
            
            try:
                industry = gisc_df.loc[ticker]['industry']
                scores = self.get_long_term_industry_rank_scores(industry)
            except:
                industry = 'None'
                scores = 'None'
                
            
            industry_score = 0
            if len(scores) != 0:
                s_rank_scores : pd.Series = self.get_long_term_industry_rank_scores(industry)
                if not s_rank_scores.empty:
                    industry_score = s_rank_scores.iloc[-1]
            else:
                print(f'can not find industry rank score from ticker: {ticker}, industry: {industry}')

            try:
                stockData = stock_datas_from_csv.get(ticker)
                atrsRank = atrs_ranking_df.loc[ticker].iloc[-1]
            except:
                continue

            volume_ma50 = stockData['Volume'].rolling(window=50).mean().iloc[-1]
            if pd.isna(volume_ma50):
                volume_ma50 = stockData['Volume'].iloc[-1]

            lower_low_3 = -1 if self.check_lower_lows_3(stockData) else 0 # bad
            higher_high_3 = 1 if self.check_higher_highs_3(stockData) else 0 # good

            below_20ma_closed = self.check_below_20ma_closed_cnt(stockData) * -1 # bad
            below_50ma_closed = self.check_below_50ma_closed_cnt(stockData) * -1 # bad

            up_more_than_adr = self.check_up_more_than_adr_cnt(stockData) # good
            down_more_than_adr = self.check_down_more_than_adr_cnt(stockData) * -1 # bad

            bullish_candle_cnt, bearish_candle_cnt = self.check_bullish_bearish_candle_count_in_n_days(stockData) # good vs bad
            more_bullish_candle = 1 if bullish_candle_cnt > bearish_candle_cnt else -1

            ma20_disparity_more_than_20ptg = self.check_20ma_dist_more_than_20ptg_cnt(stockData) * -1 # bad

            close_equal_high, close_equal_low = self.check_close_equal_high_or_low_cnt(stockData) # bad, good
            close_equal_low *= -1

            open_equal_high = self.check_OEH_cnt(stockData) * -1 # bad
            open_equal_low = self.check_OEL_cnt(stockData) # good

            oops_up_reversal = self.check_oops_up_reversal_cnt(stockData) # good

            squat = self.check_squat_cnt(stockData) * -1 # bad
            squat_recovery = self.check_squat_recovery_cnt(stockData) # good

            rs_new_high = 1 if self.check_rs_N_day_new_high_in_n_days(stockData, atrs_ranking_df, 50, 15) else 0 # good
            rs_new_low = -1 if self.check_rs_N_day_new_low_in_n_days(stockData, atrs_ranking_df, 50, 15) else 0 # bad

            pocket_pivot_cnt = self.check_pocket_pivot_cnt(stockData)


            CV_total_cnt = (lower_low_3 + higher_high_3 + below_20ma_closed + below_50ma_closed + up_more_than_adr + down_more_than_adr + more_bullish_candle +
            ma20_disparity_more_than_20ptg + close_equal_low + close_equal_high + open_equal_high + open_equal_low + oops_up_reversal + squat + squat_recovery + 
            rs_new_high + rs_new_low + pocket_pivot_cnt)


            bPocketPivot = self.check_pocket_pivot(stockData)
            bInsideBar, bDoubleInsideBar = self.check_insideBar(stockData)
            NR_x = self.check_NR_with_TrueRange(stockData)
            bWickPlay = self.check_wickplay(stockData)
            bOEL = self.check_OEL(stockData)
            bGOEL = self.check_GOEL(stockData)

            bOopsUpReversal = self.check_oops_up_reversal(stockData)
            bFailedDownsideWickBO = self.check_failed_downside_wick_BO(stockData)
            bConverging, bPower3, bPower2 = self.check_ma_converging(stockData)

            bNearEma10 = self.check_near_ma(stockData, 10, 1.5, True)
            bNearEma21 = self.check_near_ma(stockData, 21, 1.5, True)
            bNearMa50 = self.check_near_ma(stockData, 50, 1.5)

            near_ma_list = []
            if bNearEma10:
                near_ma_list.append(10)
            if bNearEma21:
                near_ma_list.append(21)
            if bNearMa50:
                near_ma_list.append(50)
            



            bNearMA = self.check_near_ma(stockData)
            ADR = stockData['ADR'].iloc[-1]

            trandingViewFormat = market + ':' + ticker + ','

            try:

                stock_info_dic[ticker] = [market, industry, industry_score, int(atrsRank), ADR, int(volume_ma50), near_ma_list, bPower3, bPocketPivot,
                                        # Volatility Contraction
                                        bInsideBar, bDoubleInsideBar, NR_x, bWickPlay,
                                        # Demand
                                        bOEL, bGOEL, bOopsUpReversal, bFailedDownsideWickBO,
                                        # C/V factors
                                        lower_low_3, higher_high_3, below_20ma_closed, below_50ma_closed, up_more_than_adr, down_more_than_adr, more_bullish_candle,
                                        ma20_disparity_more_than_20ptg, close_equal_low, close_equal_high, open_equal_high, open_equal_low, oops_up_reversal,
                                        squat, squat_recovery, rs_new_high, rs_new_low, pocket_pivot_cnt, CV_total_cnt,
                                        trandingViewFormat]
            except:
                print("Fail to convert dictionary data from cooked properties, ticker: ", ticker)
                continue

        df = pd.DataFrame.from_dict(stock_info_dic).transpose()
        columns = ['Market', 'Industry', 'Industry Score', 'RS Rank','ADR(%)', 'Volume(50avg)','Near MA list(1.5%)', 'Power of 3', 'Pocket Pivot',
                   'Inside bar', 'Double Inside bar', 'NR(x)', 'Wick Play',
                   'OEL', 'bGOEL', 'Oops up reversal', 'Failed downside wick BO',
                   'lower_low_3', 'higher_high_3', 'below_20ma_closed', 'below_50ma_closed', 'up_more_than_adr', 'down_more_than_adr', 'more_bullish_candle',
                    'ma20_disparity_more_than_20ptg', 'close_equal_low', 'close_equal_high', 'open_equal_high', 'open_equal_low', 'oops_up_reversal',
                    'squat', 'squat_recovery', 'rs_50d_new_high', 'rs_50d_new_low', 'pocket_pivot_cnt' ,'CV_total_cnt',
                    'TrandingViewFormat']
        df.columns = columns
        df.index.name = 'Symbol'


        save_path = os.path.join(FILTERED_STOCKS_FOLDER, f'{fileName}.xlsx')
        
        df.to_excel(save_path, index_label='Symbol')
        print(f'{fileName}.xlsx', 'is saved!')


        # 엑셀 조건수 서식 적용
        wb = openpyxl.load_workbook(save_path)
        sheet = wb['Sheet1']

        # 컬럼 추가되면 여기 바꿔야한다. (밑에 컬럼 색깔 정하는 범위). Inside bar의 인덱스
        min_col_start_index = 9
        NR_index = 13

        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # 연한 빨강
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # 연한 녹색

        for column in sheet.iter_cols(min_row=2, max_row=sheet.max_row, min_col=min_col_start_index, max_col=len(columns)):
            for cell in column:
                if cell.value == None:
                    continue

                # NR(x)
                if cell.column == NR_index:
                    if cell.value > 3:
                        cell.fill = green_fill
                        continue

                if cell.value == 'TRUE':
                    cell.fill = green_fill
                elif cell.value == 'FALSE':
                    cell.fill = red_fill


                if cell.value > 0:
                    cell.fill = green_fill
                elif cell.value < 0:
                    cell.fill = red_fill

        wb.save(save_path)


        cook_end_time = time.time()
        elapsedTime = cook_end_time - cook_start_time
        print('cook elapsed time: ', elapsedTime)

    
    def date_to_index(self, df : pd.DataFrame, date_str):
        """
        - return index from end. (-N)
        """
        index_position = df.index.get_loc(pd.to_datetime(date_str))
        n_day_before_index = len(df) - index_position
        return -n_day_before_index
